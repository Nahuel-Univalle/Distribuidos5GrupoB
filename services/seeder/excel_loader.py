from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

try:
    from geo_reference import gateway_safe_point, zone_center
except Exception:
    def zone_center(distrito_id: int, zona_id: int) -> tuple[float, float]:
        base_lat = -17.3895
        base_lon = -66.1568
        lat = base_lat + ((int(distrito_id) % 7) - 3) * 0.009 + ((int(zona_id) % 5) - 2) * 0.002
        lon = base_lon + ((int(distrito_id) % 6) - 3) * 0.010 + ((int(zona_id) % 7) - 3) * 0.002
        return round(lat, 6), round(lon, 6)

    def gateway_safe_point(gateway_id: int) -> tuple[float, float]:
        points = [
            (-17.389222, -66.141722),
            (-17.381000, -66.153361),
            (-17.369861, -66.176389),
            (-17.444083, -66.140694),
            (-17.420000, -66.110000),
            (-17.350000, -66.170000),
            (-17.393000, -66.157000),
            (-17.450000, -66.180000),
            (-17.389000, -66.161000),
            (-17.410000, -66.155000),
            (-17.405000, -66.135000),
            (-17.377000, -66.185000),
            (-17.355000, -66.145000),
            (-17.410000, -66.135000),
        ]
        idx = max(int(gateway_id) - 1, 0) % len(points)
        return points[idx]


SUB_ALCALDIAS: list[tuple[int, str]] = [
    (1, "TUNARI"),
    (2, "MOLLE"),
    (3, "ALEJO CALATAYUD"),
    (4, "VALLE HERMOSO"),
    (5, "ITOCTA"),
    (6, "ADELA ZAMUDIO"),
]

SUB_ALCALDIA_ID = {name.upper(): sid for sid, name in SUB_ALCALDIAS}

SUB_ALCALDIA_BY_DISTRITO: dict[int, int] = {
    1: 1,
    2: 1,
    13: 1,
    3: 2,
    4: 2,
    5: 3,
    8: 3,
    6: 4,
    7: 4,
    14: 4,
    9: 5,
    15: 5,
    10: 6,
    11: 6,
    12: 6,
}

TARIFA_HEADERS = ["R1", "R2", "R3", "R4", "C", "CE", "I", "P", "S"]

BASE_GATEWAYS: list[tuple[str, float, float]] = [
    ("LoRaWan-Teleferico", -17.389222, -66.141722),
    ("LoRaWan-ParqueVial", -17.381000, -66.153361),
    ("LoRaWan-ParqueLincoln", -17.369861, -66.176389),
    ("LoRaWan-Petrolera", -17.444083, -66.140694),
    ("LoRaWan-SurEste", -17.420000, -66.110000),
    ("LoRaWan-Norte", -17.350000, -66.170000),
    ("LoRaWan-Centro", -17.393000, -66.157000),
    ("LoRaWan-Itocta", -17.450000, -66.180000),
    ("LoRaWan-AdelaZamudio", -17.389000, -66.161000),
    ("LoRaWan-ValleHermoso", -17.410000, -66.155000),
    ("LoRaWan-AlejoCalatayud", -17.405000, -66.135000),
    ("LoRaWan-Molle", -17.377000, -66.185000),
    ("LoRaWan-Tunari", -17.355000, -66.145000),
    ("LoRaWan-LagunaAlalay", -17.410000, -66.135000),
]


def target_gateways() -> int:
    return int(os.getenv("SEED_TARGET_GATEWAYS", "14"))


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def norm_key(value: Any) -> str:
    text = clean_text(value).upper()
    replacements = {
        "Á": "A",
        "É": "E",
        "Í": "I",
        "Ó": "O",
        "Ú": "U",
        "Ñ": "N",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        found = re.search(r"\d+", clean_text(value))
        return int(found.group(0)) if found else None


def _as_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    text = clean_text(value)
    text = text.replace("Bs.", "").replace("Bs", "").replace("$us", "").replace("USD", "")
    text = text.replace(" ", "").replace(",", ".")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _parse_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = clean_text(value)
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%m/%d/%y %H:%M",
        "%d/%m/%y %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def gateway_id_from_name(raw_name: Any) -> int:
    text = norm_key(raw_name)
    if not text:
        return 1

    for idx, (name, _lat, _lon) in enumerate(BASE_GATEWAYS, start=1):
        if norm_key(name) in text or text in norm_key(name):
            return ((idx - 1) % target_gateways()) + 1

    number = _as_int(text)
    if number:
        return ((number - 1) % target_gateways()) + 1

    return 1


def gateway_pool_for(gateway_id: int) -> list[int]:
    target = max(target_gateways(), 1)
    gid = ((int(gateway_id or 1) - 1) % target) + 1
    return [gid]


@dataclass(slots=True)
class Distrito:
    distrito_id: int
    sub_alcaldia_id: int
    nombre: str
    habitantes: int = 0


@dataclass(slots=True)
class Zona:
    distrito_id: int
    zona_id: int
    nombre: str
    gateway_id: int
    habitantes: int
    counts: dict[str, int] = field(default_factory=dict)
    centro_lat: float = -17.39
    centro_lon: float = -66.15

    @property
    def total_medidores(self) -> int:
        return sum(int(v or 0) for v in self.counts.values())


@dataclass(slots=True)
class ModeloMedidor:
    modelo_id: int
    marca: str
    modelo: str
    conectividad: str
    aplicacion: str


@dataclass(slots=True)
class TarifaCat:
    categoria: str
    alias: str
    fijo_m3: Decimal
    usd_mes: Decimal
    r_13_25: Decimal
    r_26_50: Decimal
    r_51_75: Decimal
    r_76_100: Decimal
    r_101_150: Decimal
    r_mas_151: Decimal
    descripcion: str


@dataclass(slots=True)
class TipoInfra:
    tipo_id: int
    descripcion: str


@dataclass(slots=True)
class UnidadEducativa:
    codigo: str
    nombre: str
    distrito_txt: str
    zona_txt: str
    direccion: str
    educacion: str


@dataclass(slots=True)
class InfraestructuraTemplate:
    numero_catastro: str
    propietario: str
    ci: str
    direccion: str
    zona: str
    distrito_id: int | None
    manzano: int | None
    lote: int | None
    superficie_terreno: int | None
    area_construida: int | None
    uso_suelo: str
    matricula_ddrr: str
    valor_catastral: Decimal
    impuesto_anual: Decimal


@dataclass(slots=True)
class ContratoTemplate:
    numero_catastro: str
    titular: str
    ci_titular: str
    categoria: str
    subcategoria: str
    medidor_iot: str
    fecha_contrato: datetime | None
    estado_contrato: str
    diametro_conexion: str
    tipo_servicio: str


@dataclass(slots=True)
class MedidorTemplate:
    medidor_iot: str
    fecha_instalacion: datetime | None
    fecha_desinstalacion: datetime | None
    estado: str
    tipo_medidor_id: int | None


@dataclass(slots=True)
class LecturaTemplate:
    medidor_iot: str
    lectura_anterior: int
    lectura_actual: int
    fecha_hora: datetime | None
    radiobase: int | None
    fecha_pago: datetime | None


def load_workbook(path: str | Path) -> openpyxl.Workbook:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel no encontrado: {p}")
    logger.info(f"Cargando Excel: {p}")
    return openpyxl.load_workbook(p, data_only=True, read_only=False)


def _sheet(wb: openpyxl.Workbook, *names: str):
    lower = {sheet.lower(): sheet for sheet in wb.sheetnames}
    for name in names:
        found = lower.get(name.lower())
        if found:
            return wb[found]
    raise KeyError(f"No se encontró ninguna hoja: {names}. Disponibles: {wb.sheetnames}")


def load_distritos_zonas(wb: openpyxl.Workbook) -> tuple[list[Distrito], list[Zona]]:
    ws = _sheet(wb, "Distritos")
    header_row = 2
    headers = [clean_text(cell.value) for cell in ws[header_row]]

    tariff_cols: dict[str, int] = {}
    for idx, header in enumerate(headers):
        header_upper = header.upper()
        if header_upper in TARIFA_HEADERS:
            tariff_cols[header_upper] = idx

    missing = set(TARIFA_HEADERS) - set(tariff_cols)
    if missing:
        raise ValueError(f"La hoja Distritos no tiene todas las tarifas {TARIFA_HEADERS}. Faltan={sorted(missing)}")

    col_sub = 0
    col_dist = 1
    col_zona = 2
    col_nombre_zona = 3
    col_gateway = 4
    col_zone_pop = 6
    col_sub_hab = 7
    col_total = next((idx for idx, header in enumerate(headers) if header.upper() == "TOTAL"), 17)

    distritos: dict[int, Distrito] = {}
    zonas: list[Zona] = []
    cur_sub = ""
    cur_dist: int | None = None
    cur_habitantes = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row or all(value is None for value in row):
            continue

        if len(row) > col_sub and row[col_sub] is not None:
            cur_sub = clean_text(row[col_sub]).upper()

        if len(row) > col_dist and row[col_dist] is not None:
            cur_dist = _as_int(row[col_dist])
            cur_habitantes = _as_int(row[col_sub_hab] if len(row) > col_sub_hab else None) or cur_habitantes
            if cur_dist is not None and cur_dist not in distritos:
                distritos[cur_dist] = Distrito(
                    distrito_id=cur_dist,
                    sub_alcaldia_id=SUB_ALCALDIA_BY_DISTRITO.get(cur_dist, SUB_ALCALDIA_ID.get(cur_sub.replace("\n", " "), 1)),
                    nombre=f"DISTRITO {cur_dist}",
                    habitantes=cur_habitantes,
                )
        elif cur_dist is not None and len(row) > col_sub_hab and row[col_sub_hab] is not None and distritos[cur_dist].habitantes == 0:
            distritos[cur_dist].habitantes = _as_int(row[col_sub_hab]) or 0

        zona_id = _as_int(row[col_zona] if len(row) > col_zona else None)
        zona_nombre = clean_text(row[col_nombre_zona] if len(row) > col_nombre_zona else None)
        if cur_dist is None or zona_id is None or not zona_nombre:
            continue

        counts = {cat: _as_int(row[col] if col < len(row) else None) or 0 for cat, col in tariff_cols.items()}
        total_col = _as_int(row[col_total] if col_total < len(row) else None) or 0
        zona_habitantes = _as_int(row[col_zone_pop] if len(row) > col_zone_pop else None) or 0

        if not any(counts.values()) and total_col == 0:
            continue

        if total_col and sum(counts.values()) != total_col:
            logger.warning(
                f"Fila {row_idx}: suma tarifas={sum(counts.values())} != Total={total_col} "
                f"en D{cur_dist}/Z{zona_id} {zona_nombre}"
            )

        gw_id = gateway_id_from_name(row[col_gateway] if len(row) > col_gateway else None)
        centro = zone_center(cur_dist, zona_id)
        zonas.append(
            Zona(
                distrito_id=cur_dist,
                zona_id=zona_id,
                nombre=zona_nombre,
                gateway_id=gw_id,
                habitantes=zona_habitantes,
                counts={cat: counts.get(cat, 0) for cat in TARIFA_HEADERS},
                centro_lat=centro[0],
                centro_lon=centro[1],
            )
        )

    for dist in distritos.values():
        zonas_distrito = [zona for zona in zonas if zona.distrito_id == dist.distrito_id]
        suma_zonas = sum(zona.habitantes for zona in zonas_distrito)
        if suma_zonas > 0:
            dist.habitantes = suma_zonas
        elif dist.habitantes and zonas_distrito:
            total_medidores = sum(zona.total_medidores for zona in zonas_distrito) or 1
            for zona in zonas_distrito:
                zona.habitantes = int(dist.habitantes * zona.total_medidores / total_medidores)

    total_base = sum(zona.total_medidores for zona in zonas)
    logger.info(f"Distritos cargados: {len(distritos)} | Zonas: {len(zonas)} | Total base={total_base:,}")
    if total_base not in {80000, 100000, 120000}:
        logger.warning(f"La hoja Distritos suma {total_base:,}. Se usará como distribución territorial de apoyo.")

    return sorted(distritos.values(), key=lambda item: item.distrito_id), zonas


def load_tarifas(wb: openpyxl.Workbook) -> list[TarifaCat]:
    ws = _sheet(wb, "Tarifario")
    out: list[TarifaCat] = []
    cur_alias = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or all(value is None for value in row):
            continue
        alias_txt, cat, fijo, usd, r1, r2, r3, r4, r5, r6, desc = (list(row) + [None] * 11)[:11]
        if alias_txt:
            cur_alias = clean_text(alias_txt)
        cat_txt = clean_text(cat).upper()
        if not cat_txt or cat_txt not in TARIFA_HEADERS:
            continue
        out.append(
            TarifaCat(
                cat_txt,
                cur_alias,
                _as_decimal(fijo),
                _as_decimal(usd),
                _as_decimal(r1),
                _as_decimal(r2),
                _as_decimal(r3),
                _as_decimal(r4),
                _as_decimal(r5),
                _as_decimal(r6),
                clean_text(desc),
            )
        )

    found = {tarifa.categoria for tarifa in out}
    missing = set(TARIFA_HEADERS) - found
    if missing:
        raise ValueError(f"Faltan tarifas en Tarifario: {sorted(missing)}")
    logger.info(f"Tarifas: {len(out)} categorías")
    return out


def load_modelos(wb: openpyxl.Workbook) -> list[ModeloMedidor]:
    ws = _sheet(wb, "ModeloMedidores")
    out: list[ModeloMedidor] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        modelo_id = _as_int(row[0])
        if modelo_id is None:
            continue
        out.append(
            ModeloMedidor(
                modelo_id,
                clean_text(row[1] if len(row) > 1 else ""),
                clean_text(row[2] if len(row) > 2 else ""),
                clean_text(row[3] if len(row) > 3 else ""),
                clean_text(row[4] if len(row) > 4 else ""),
            )
        )
    logger.info(f"Modelos medidor: {len(out)}")
    return out


def load_errores(wb: openpyxl.Workbook) -> list[tuple[int, str]]:
    ws = _sheet(wb, "ErroresIOT")
    out: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = _as_int(row[0])
        if code is None:
            continue
        out.append((code, clean_text(row[1] if len(row) > 1 else "")))
    logger.info(f"Errores IoT: {len(out)}")
    return out


def load_tipos_infra(wb: openpyxl.Workbook) -> list[TipoInfra]:
    base = [
        "Educativo",
        "Salud",
        "Asilo / Convento / Iglesia",
        "Beneficencia",
        "Área verde / Parque",
        "Centro comunal / Cultural",
        "Infraestructura pública / Hidrante",
        "Terreno baldío",
        "Casa abandonada",
        "Edificio",
        "Condominio",
        "Residencial",
        "Comercial",
        "Comercial Especial",
        "Industrial",
        "Mixto",
    ]
    usos: list[str] = []
    try:
        ws = _sheet(wb, "Infraestructura")
        headers = [norm_key(cell.value) for cell in ws[1]]
        col_uso = headers.index("USO_SUELO") if "USO_SUELO" in headers else None
        if col_uso is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = clean_text(row[col_uso] if col_uso < len(row) else None)
                if val and val not in usos:
                    usos.append(val)
    except Exception:
        pass

    merged: list[str] = []
    for value in usos + base:
        if value and value not in merged:
            merged.append(value)

    out = [TipoInfra(index + 1, desc) for index, desc in enumerate(merged)]
    logger.info(f"Tipos infraestructura: {len(out)}")
    return out


def load_unidades_educativas(wb: openpyxl.Workbook) -> list[UnidadEducativa]:
    ws = _sheet(wb, "UnidadesEducativas")
    out: list[UnidadEducativa] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 9 or row[1] is None:
            continue
        out.append(
            UnidadEducativa(
                codigo=clean_text(row[1]),
                nombre=clean_text(row[2]),
                distrito_txt=clean_text(row[0]),
                zona_txt=clean_text(row[7]),
                direccion=clean_text(row[8]),
                educacion=clean_text(row[3]),
            )
        )
    logger.info(f"Unidades educativas: {len(out)}")
    return out


def load_infraestructura_templates(wb: openpyxl.Workbook) -> list[InfraestructuraTemplate]:
    ws = _sheet(wb, "Infraestructura")
    out: list[InfraestructuraTemplate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out.append(
            InfraestructuraTemplate(
                numero_catastro=clean_text(row[0]),
                propietario=clean_text(row[1] if len(row) > 1 else ""),
                ci=clean_text(row[2] if len(row) > 2 else ""),
                direccion=clean_text(row[3] if len(row) > 3 else ""),
                zona=clean_text(row[4] if len(row) > 4 else ""),
                distrito_id=_as_int(row[5] if len(row) > 5 else None),
                manzano=_as_int(row[6] if len(row) > 6 else None),
                lote=_as_int(row[7] if len(row) > 7 else None),
                superficie_terreno=_as_int(row[8] if len(row) > 8 else None),
                area_construida=_as_int(row[9] if len(row) > 9 else None),
                uso_suelo=clean_text(row[10] if len(row) > 10 else ""),
                matricula_ddrr=clean_text(row[11] if len(row) > 11 else ""),
                valor_catastral=_as_decimal(row[12] if len(row) > 12 else None),
                impuesto_anual=_as_decimal(row[13] if len(row) > 13 else None),
            )
        )
    logger.info(f"Plantillas infraestructura/catastro: {len(out)}")
    return out


def load_contratos_templates(wb: openpyxl.Workbook) -> list[ContratoTemplate]:
    ws = _sheet(wb, "Contratos")
    out: list[ContratoTemplate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out.append(
            ContratoTemplate(
                numero_catastro=clean_text(row[0]),
                titular=clean_text(row[1] if len(row) > 1 else ""),
                ci_titular=clean_text(row[2] if len(row) > 2 else ""),
                categoria=clean_text(row[3] if len(row) > 3 else ""),
                subcategoria=clean_text(row[4] if len(row) > 4 else "").upper(),
                medidor_iot=clean_text(row[5] if len(row) > 5 else ""),
                fecha_contrato=_parse_date(row[6] if len(row) > 6 else None),
                estado_contrato=clean_text(row[7] if len(row) > 7 else "").upper(),
                diametro_conexion=clean_text(row[8] if len(row) > 8 else ""),
                tipo_servicio=clean_text(row[9] if len(row) > 9 else ""),
            )
        )
    logger.info(f"Plantillas contratos: {len(out)}")
    return out


def load_medidores_templates(wb: openpyxl.Workbook) -> list[MedidorTemplate]:
    ws = _sheet(wb, "Medidores")
    out: list[MedidorTemplate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out.append(
            MedidorTemplate(
                medidor_iot=clean_text(row[0]),
                fecha_instalacion=_parse_date(row[1] if len(row) > 1 else None),
                fecha_desinstalacion=_parse_date(row[2] if len(row) > 2 else None),
                estado=clean_text(row[3] if len(row) > 3 else ""),
                tipo_medidor_id=_as_int(row[4] if len(row) > 4 else None),
            )
        )
    logger.info(f"Plantillas medidores: {len(out)}")
    return out


def load_lecturas_templates(wb: openpyxl.Workbook) -> list[LecturaTemplate]:
    ws = _sheet(wb, "Lecturas")
    out: list[LecturaTemplate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out.append(
            LecturaTemplate(
                medidor_iot=clean_text(row[0]),
                lectura_anterior=_as_int(row[1] if len(row) > 1 else None) or 0,
                lectura_actual=_as_int(row[2] if len(row) > 2 else None) or 0,
                fecha_hora=_parse_date(row[3] if len(row) > 3 else None),
                radiobase=_as_int(row[4] if len(row) > 4 else None),
                fecha_pago=_parse_date(row[5] if len(row) > 5 else None),
            )
        )
    logger.info(f"Plantillas lecturas: {len(out)}")
    return out


def make_catastro_number(distrito_id: int, zona_id: int, manzano: int, lote: int, subdivision: int = 0) -> str:
    return f"{int(distrito_id):02d}-{int(zona_id):02d}-{int(manzano) % 1000:03d}-{int(lote) % 10000:04d}-{int(subdivision) % 1000:03d}"


def gateways() -> list[tuple[int, str, float, float]]:
    target = target_gateways()
    out: list[tuple[int, str, float, float]] = []
    for index in range(target):
        name, lat, lon = BASE_GATEWAYS[index % len(BASE_GATEWAYS)]
        if index >= len(BASE_GATEWAYS):
            lat += (index // len(BASE_GATEWAYS)) * 0.002
            lon -= (index // len(BASE_GATEWAYS)) * 0.002
        safe_lat, safe_lon = gateway_safe_point(index + 1)
        out.append((index + 1, name, safe_lat or round(lat, 6), safe_lon or round(lon, 6)))
    return out
